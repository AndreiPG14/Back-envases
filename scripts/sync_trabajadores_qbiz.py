#!/usr/bin/env python3
"""
Script para sincronizar trabajadores desde OneDrive QBIZ a Supabase
Se ejecuta automáticamente cada lunes a las 6am UTC via GitHub Actions
"""

import os
import sys
import requests
import openpyxl
from io import BytesIO
from datetime import datetime
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

API_URL = os.getenv('API_URL')
ONEDRIVE_SHARE_URL = os.getenv('ONEDRIVE_SHARE_URL')

if not API_URL:
    print("❌ Error: API_URL no configurada")
    sys.exit(1)

if not ONEDRIVE_SHARE_URL:
    print("❌ Error: ONEDRIVE_SHARE_URL no configurada")
    sys.exit(1)

def descargar_excel() -> BytesIO:
    """Descargar Excel QBIZ desde OneDrive"""
    print("📥 Descargando Excel QBIZ desde OneDrive...")

    try:
        session = requests.Session()
        response = session.get(
            ONEDRIVE_SHARE_URL,
            params={'download': '1'},
            allow_redirects=True,
            headers={'User-Agent': 'Mozilla/5.0'},
            timeout=30
        )

        if response.status_code != 200:
            raise Exception(f"Status code: {response.status_code}")

        print(f"✅ Excel descargado ({len(response.content)} bytes)")
        return BytesIO(response.content)
    except Exception as e:
        print(f"❌ Error descargando Excel: {e}")
        raise

def leer_filas(excel_bytes: BytesIO) -> list:
    """Leer filas del Excel"""
    print("📖 Leyendo datos del Excel...")

    try:
        wb = openpyxl.load_workbook(excel_bytes, data_only=True)
        ws = wb.active

        # Obtener headers de la primera fila
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        print(f"📋 Columnas encontradas: {len(headers)}")

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            fila = {}
            for i, valor in enumerate(row):
                if i < len(headers) and headers[i]:
                    # Formatear fechas
                    if hasattr(valor, 'strftime'):
                        valor = valor.strftime('%d/%m/%Y')
                    fila[headers[i]] = str(valor).strip() if valor is not None else ''

            # Solo agregar si tiene DNI válido
            if fila.get('DNI') and fila.get('DNI') != 'None' and fila.get('DNI').strip():
                rows.append(fila)

        print(f"✅ {len(rows)} filas con datos válidos")

        # Mostrar empresas distintas
        empresas = set(r.get('EMPRESA', '') for r in rows if r.get('EMPRESA'))
        print(f"🏢 Empresas: {', '.join(empresas) if empresas else 'N/A'}")

        return rows
    except Exception as e:
        print(f"❌ Error leyendo Excel: {e}")
        raise

def enviar_en_lotes(rows: list, lote_size: int = 200) -> dict:
    """Enviar datos a la API en lotes"""
    total_insertados = 0
    total_omitidos = 0
    total_lotes = (len(rows) + lote_size - 1) // lote_size

    for i in range(0, len(rows), lote_size):
        lote = rows[i:i + lote_size]
        lote_num = (i // lote_size) + 1

        print(f"📤 Lote {lote_num}/{total_lotes} ({len(lote)} filas)...")

        try:
            response = requests.post(
                f'{API_URL}/api/trabajadores-qbiz/sync',
                json={'rows': lote},
                headers={'Content-Type': 'application/json'},
                timeout=120
            )

            if response.status_code not in [200, 201]:
                print(f"❌ Error lote {lote_num}: {response.status_code}")
                print(f"   Reintentando fila a fila...")

                # Reintentar fila a fila
                for idx, fila in enumerate(lote):
                    try:
                        r2 = requests.post(
                            f'{API_URL}/api/trabajadores-qbiz/sync',
                            json={'rows': [fila]},
                            headers={'Content-Type': 'application/json'},
                            timeout=30
                        )

                        if r2.status_code in [200, 201]:
                            resultado = r2.json()
                            total_insertados += resultado.get('data', {}).get('insertados', 0)
                            total_omitidos += resultado.get('data', {}).get('omitidos', 0)
                        else:
                            print(f"   ❌ DNI={fila.get('DNI')}: {r2.text[:100]}")
                    except Exception as e2:
                        print(f"   ❌ DNI={fila.get('DNI')}: {e2}")
                continue

            resultado = response.json()
            insertados = resultado.get('data', {}).get('insertados', 0)
            omitidos = resultado.get('data', {}).get('omitidos', 0)
            total_insertados += insertados
            total_omitidos += omitidos
            print(f"   ✅ {insertados} insertados, {omitidos} omitidos")

        except Exception as e:
            print(f"❌ Error lote {lote_num}: {e}")
            continue

    return {'insertados': total_insertados, 'omitidos': total_omitidos}

def main():
    """Flujo principal"""
    print("\n" + "="*60)
    print("🔄 Sync Trabajadores QBIZ → Supabase")
    print(f"⏰ {datetime.now().isoformat()}")
    print("="*60 + "\n")

    try:
        # Paso 1: Descargar Excel
        print("[1/3] Descargando Excel...")
        excel_bytes = descargar_excel()

        # Paso 2: Leer datos
        print("[2/3] Extrayendo datos...")
        rows = leer_filas(excel_bytes)

        if not rows:
            print("⚠️  No hay filas válidas para sincronizar")
            return 0

        # Paso 3: Enviar a API
        print("[3/3] Sincronizando a API...")
        resultado = enviar_en_lotes(rows, lote_size=200)

        # Resultado final
        print("\n" + "="*60)
        print(f"✅ Total insertados/actualizados: {resultado['insertados']}")
        print(f"⏭️  Total omitidos: {resultado['omitidos']}")
        print("="*60 + "\n")

        return 0

    except Exception as e:
        print(f"\n❌ Error fatal: {e}")
        print("="*60 + "\n")
        return 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
